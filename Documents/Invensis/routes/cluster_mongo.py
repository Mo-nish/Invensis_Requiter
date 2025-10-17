from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from models_mongo import User, Candidate, ActivityLog, users_collection
from datetime import datetime, timedelta
from email_service import send_candidate_assignment_email
from bson import ObjectId

cluster_bp = Blueprint('cluster', __name__)

def cluster_required(f):
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Please login first.', 'error')
            return redirect(url_for('login'))
        if current_user.role != 'cluster':
            flash('Access denied. Cluster privileges required.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@cluster_bp.route('/dashboard')
@cluster_required
def dashboard():
    # Get filter parameters
    status_filter = request.args.get('status')
    hr_filter = request.args.get('hr')
    manager_filter = request.args.get('manager')
    cluster_filter = request.args.get('cluster')
    date_filter = request.args.get('date')
    search_filter = request.args.get('search')
    
    # Get all candidates for analytics
    from models_mongo import candidates_collection, users_collection, activity_logs_collection
    
    # Build query based on filters
    query = {}
    if status_filter:
        query['status'] = status_filter
    if hr_filter:
        # Find HR user by name to get their email
        hr_user = users_collection.find_one({'role': 'hr_role', 'name': hr_filter, 'is_active': True})
        if hr_user:
            query['assigned_by'] = hr_user['email']
    if manager_filter:
        # Find Manager user by name to get their email
        manager_user = users_collection.find_one({'role': 'manager', 'name': manager_filter, 'is_active': True})
        if manager_user:
            query['manager_email'] = manager_user['email']
    if cluster_filter:
        # Find users in the selected cluster
        cluster_users = list(users_collection.find({'cluster': cluster_filter, 'is_active': True}))
        if cluster_users:
            cluster_emails = [user['email'] for user in cluster_users]
            # Filter candidates by cluster users (either assigned_by or manager_email)
            if '$or' in query:
                # If we already have $or for search, we need to combine with AND logic
                existing_or = query['$or']
                query['$and'] = [
                    {'$or': existing_or},
                    {'$or': [
                        {'assigned_by': {'$in': cluster_emails}},
                        {'manager_email': {'$in': cluster_emails}}
                    ]}
                ]
                del query['$or']
            else:
                query['$or'] = [
                    {'assigned_by': {'$in': cluster_emails}},
                    {'manager_email': {'$in': cluster_emails}}
                ]
    if date_filter:
        # Convert date string to datetime for comparison
        from datetime import datetime
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d')
            query['created_at'] = {
                '$gte': filter_date,
                '$lt': filter_date.replace(hour=23, minute=59, second=59)
            }
        except ValueError:
            pass  # Invalid date format, ignore filter
    if search_filter:
        # Search in name, email, and phone fields
        import re
        search_regex = re.compile(search_filter, re.IGNORECASE)
        query['$or'] = [
            {'first_name': search_regex},
            {'last_name': search_regex},
            {'email': search_regex},
            {'phone': search_regex}
        ]
    
    # Get filtered candidates
    all_candidates_raw = list(candidates_collection.find(query))
    
    # Convert to Candidate objects for proper data handling
    all_candidates = []
    for data in all_candidates_raw:
        try:
            candidate = Candidate.from_dict(data)
            all_candidates.append(candidate)
        except Exception as e:
            print(f"Error converting candidate data: {e}")
            continue
    
    # Calculate statistics based on filtered data
    total_candidates = len(all_candidates)
    selected_candidates = len([c for c in all_candidates if c.status == 'Selected'])
    not_selected_candidates = len([c for c in all_candidates if c.status == 'Not Selected'])
    # Reassigned candidates are those with reassigned_by_manager field, not a specific status
    reassigned_candidates = len([c for c in all_candidates if c.reassigned_by_manager is not None])

    
    # Calculate average ratings
    average_hr_rating = 0
    average_manager_rating = 0
    
    candidates_with_hr_rating = [c for c in all_candidates if c.overall_rating and c.overall_rating > 0]
    if candidates_with_hr_rating:
        average_hr_rating = sum(c.overall_rating for c in candidates_with_hr_rating) / len(candidates_with_hr_rating)
    
    candidates_with_manager_rating = [c for c in all_candidates if c.manager_overall_rating and c.manager_overall_rating > 0]
    if candidates_with_manager_rating:
        average_manager_rating = sum(c.manager_overall_rating for c in candidates_with_manager_rating) / len(candidates_with_manager_rating)
    
    # Get recent candidates for display (apply same filters)
    recent_candidates_raw = list(candidates_collection.find(query).sort('created_at', -1).limit(10))
    
    # Convert recent candidates to Candidate objects
    recent_candidates = []
    for data in recent_candidates_raw:
        try:
            candidate = Candidate.from_dict(data)
            recent_candidates.append(candidate)
        except Exception as e:
            print(f"Error converting recent candidate data: {e}")
            continue
    
    # Get HR names for filter
    hr_users = list(users_collection.find({'role': 'hr_role', 'is_active': True}))
    hr_names = []
    for hr in hr_users:
        hr_names.append(hr['name'])
    
    # Get Manager names for filter
    manager_users = list(users_collection.find({'role': 'manager', 'is_active': True}))
    manager_names = []
    for manager in manager_users:
        manager_names.append(manager['name'])
    
    # Get unique clusters for filter (from users who have cluster info)
    cluster_users = list(users_collection.find({'cluster': {'$exists': True, '$ne': None}}))
    cluster_names = []
    for user in cluster_users:
        if user.get('cluster') and user['cluster'] not in cluster_names:
            cluster_names.append(user['cluster'])
    
    # If no clusters found, provide default options
    if not cluster_names:
        cluster_names = ['Tech Cluster', 'Sales Cluster', 'Marketing Cluster', 'Operations Cluster']
    
    return render_template('cluster/dashboard.html',
                         total_candidates=total_candidates,
                         selected_candidates=selected_candidates,
                         not_selected_candidates=not_selected_candidates,
                         reassigned_candidates=reassigned_candidates,

                         recent_candidates=recent_candidates,
                         hr_names=hr_names,
                         manager_names=manager_names,
                         cluster_names=cluster_names,
                         average_hr_rating=average_hr_rating,
                         average_manager_rating=average_manager_rating,
                         candidates=all_candidates)

@cluster_bp.route('/debug-candidates')
@cluster_required
def debug_candidates():
    """Debug route to see what candidates exist in the database"""
    from models_mongo import candidates_collection
    
    # Get all candidates
    all_candidates = list(candidates_collection.find({}))
    
    # Get unique status values
    statuses = candidates_collection.distinct('status')
    
    # Get unique assigned_by values
    assigned_by = candidates_collection.distinct('assigned_by')
    
    # Get unique manager_email values
    manager_emails = candidates_collection.distinct('manager_email')
    
    return jsonify({
        'total_candidates': len(all_candidates),
        'statuses': statuses,
        'assigned_by': assigned_by,
        'manager_emails': manager_emails,
        'sample_candidate': all_candidates[0] if all_candidates else None
    })

@cluster_bp.route('/candidates')
@cluster_required
def candidates():
    from models_mongo import candidates_collection
    candidates_data = list(candidates_collection.find().sort('created_at', -1))
    
    # Convert to Candidate objects for proper data handling
    candidates = []
    for data in candidates_data:
        try:
            candidate = Candidate.from_dict(data)
            candidates.append(candidate)
        except Exception as e:
            print(f"Error converting candidate data: {e}")
            continue
    
    return render_template('cluster/candidates.html', candidates=candidates)

@cluster_bp.route('/assign_candidate', methods=['POST'])
@cluster_required
def assign_candidate():
    candidate_id = request.form['candidate_id']
    manager_email = request.form['manager_email']
    interview_time = request.form.get('interview_time')
    
    # Validate manager email format
    if not manager_email or '@' not in manager_email:
        return jsonify({'success': False, 'message': 'Please enter a valid manager email address'})
    
    candidate = Candidate.find_by_id(candidate_id)
    if candidate:
        # Convert interview_time string to datetime object
        interview_datetime = None
        if interview_time:
            try:
                interview_datetime = datetime.fromisoformat(interview_time.replace('Z', '+00:00'))
            except ValueError:
                return jsonify({'success': False, 'message': 'Please enter a valid interview date and time'})
        
        candidate.manager_email = manager_email
        candidate.scheduled_date = interview_datetime
        candidate.status = 'Assigned'
        candidate.assigned_to_manager = True
        candidate.save()
        
        # Send email to manager
        try:
            send_candidate_assignment_email(manager_email, candidate, interview_datetime)
        except Exception as e:
            flash('Email notification failed', 'warning')
        
        # Log activity
        activity = ActivityLog(
            user_email=current_user.email,
            action='Assigned candidate',
            target_email=manager_email,
            details=f'Cluster assigned candidate {candidate.first_name} {candidate.last_name} to {manager_email} for interview on {interview_datetime.strftime("%Y-%m-%d %H:%M") if interview_datetime else "TBD"}'
        )
        activity.save()
        
        return jsonify({'success': True, 'message': 'Candidate assigned successfully'})
    else:
        return jsonify({'success': False, 'message': 'Candidate not found'})

@cluster_bp.route('/candidate/<candidate_id>')
@cluster_required
def candidate_details(candidate_id):
    candidate = Candidate.find_by_id(candidate_id)
    if candidate:
        return render_template('cluster/candidate_details.html', candidate=candidate)
    else:
        flash('Candidate not found', 'error')
        return redirect(url_for('cluster.dashboard'))

@cluster_bp.route('/analytics')
@cluster_required
def analytics():
    from models_mongo import candidates_collection, users_collection
    
    # Get all data for analytics
    all_candidates = list(candidates_collection.find())
    all_users = list(users_collection.find({'is_active': True}))
    
    # Status distribution
    status_counts = {}
    for candidate in all_candidates:
        status = candidate.get('status', 'Pending')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    # Monthly trends
    monthly_data = {}
    for candidate in all_candidates:
        month = candidate.get('created_at', datetime.now()).strftime('%Y-%m')
        monthly_data[month] = monthly_data.get(month, 0) + 1
    
    # Manager performance
    manager_stats = {}
    for user in all_users:
        if user['role'] == 'manager':
            manager_candidates = [c for c in all_candidates if c.get('manager_email') == user['email']]
            manager_stats[user['name']] = {
                'total': len(manager_candidates),
                'shortlisted': len([c for c in manager_candidates if c.get('status') == 'Shortlisted']),
                'rejected': len([c for c in manager_candidates if c.get('status') == 'Rejected']),
                'pending': len([c for c in manager_candidates if c.get('status') in ['Pending', 'Assigned']])
            }
    
    return render_template('cluster/analytics.html',
                         status_counts=status_counts,
                         monthly_data=monthly_data,
                         manager_stats=manager_stats)

@cluster_bp.route('/reports')
@cluster_required
def reports():
    from models_mongo import candidates_collection, activity_logs_collection
    
    # Generate various reports
    all_candidates = list(candidates_collection.find())
    all_activities = list(activity_logs_collection.find())
    
    # Recent activity report
    recent_activities = list(activity_logs_collection.find().sort('timestamp', -1).limit(50))
    
    # Status summary report
    status_summary = {}
    for candidate in all_candidates:
        status = candidate.get('status', 'Pending')
        status_summary[status] = status_summary.get(status, 0) + 1
    
    return render_template('cluster/reports.html',
                         recent_activities=recent_activities,
                         status_summary=status_summary,
                         total_candidates=len(all_candidates))

@cluster_bp.route('/export_data')
@cluster_required
def export_data():
    from flask import Response, send_file
    import csv
    import io
    import os
    from datetime import datetime
    
    # Import openpyxl for Excel export functionality
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill  # type: ignore
    except ImportError:
        # Fallback if openpyxl is not available
        return jsonify({'error': 'Excel export not available - openpyxl not installed'}), 400
    
    # Get export format
    export_format = request.args.get('format', 'csv')
    
    # Get filter parameters (same as dashboard)
    status_filter = request.args.get('status')
    hr_filter = request.args.get('hr')
    date_filter = request.args.get('date')
    search_filter = request.args.get('search')
    
    # Build query based on filters (same logic as dashboard)
    query = {}
    if status_filter:
        query['status'] = status_filter
    if hr_filter:
        # Find HR user by name to get their email
        hr_user = users_collection.find_one({'role': 'hr_role', 'name': hr_filter, 'is_active': True})
        if hr_user:
            query['assigned_by'] = hr_user['email']
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d')
            query['created_at'] = {
                '$gte': filter_date,
                '$lt': filter_date.replace(hour=23, minute=59, second=59)
            }
        except ValueError:
            pass
    if search_filter:
        import re
        search_regex = re.compile(search_filter, re.IGNORECASE)
        query['$or'] = [
            {'first_name': search_regex},
            {'last_name': search_regex},
            {'email': search_regex},
            {'phone': search_regex}
        ]
    
    # Get filtered candidates
    from models_mongo import candidates_collection
    candidates = list(candidates_collection.find(query))
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    if export_format == 'excel':
        # Create Excel file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Cluster Export"
        
        # Define headers
        headers = [
            'Reference ID', 'First Name', 'Last Name', 'Email', 'Phone', 
            'Status', 'Assigned HR', 'Manager Email', 'Created Date', 
            'Scheduled Date', 'Experience', 'Skills'
        ]
        
        # Style for headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        
        # Write data rows
        for row, candidate in enumerate(candidates, 2):
            created_date = candidate.get('created_at')
            if created_date:
                if isinstance(created_date, datetime):
                    created_date = created_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    created_date = str(created_date)
            else:
                created_date = ''
                
            scheduled_date = candidate.get('scheduled_date')
            if scheduled_date:
                if isinstance(scheduled_date, datetime):
                    scheduled_date = scheduled_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    scheduled_date = str(scheduled_date)
            else:
                scheduled_date = ''
            
            row_data = [
                candidate.get('reference_id', ''),
                candidate.get('first_name', ''),
                candidate.get('last_name', ''),
                candidate.get('email', ''),
                candidate.get('phone', ''),
                candidate.get('status', ''),
                candidate.get('assigned_by', ''),
                candidate.get('manager_email', ''),
                created_date,
                scheduled_date,
                candidate.get('experience', ''),
                candidate.get('skills', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        filename = f'cluster_export_{timestamp}.xlsx'
        filepath = f'/tmp/{filename}'
        wb.save(filepath)
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    else:
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow([
            'Reference ID', 'First Name', 'Last Name', 'Email', 'Phone', 
            'Status', 'Assigned HR', 'Manager Email', 'Created Date', 
            'Scheduled Date', 'Experience', 'Skills'
        ])
        
        # Write data rows
        for candidate in candidates:
            created_date = candidate.get('created_at')
            if created_date:
                if isinstance(created_date, datetime):
                    created_date = created_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    created_date = str(created_date)
            else:
                created_date = ''
                
            scheduled_date = candidate.get('scheduled_date')
            if scheduled_date:
                if isinstance(scheduled_date, datetime):
                    scheduled_date = scheduled_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    scheduled_date = str(scheduled_date)
            else:
                scheduled_date = ''
            
            writer.writerow([
                candidate.get('reference_id', ''),
                candidate.get('first_name', ''),
                candidate.get('last_name', ''),
                candidate.get('email', ''),
                candidate.get('phone', ''),
                candidate.get('status', ''),
                candidate.get('assigned_by', ''),
                candidate.get('manager_email', ''),
                created_date,
                scheduled_date,
                candidate.get('experience', ''),
                candidate.get('skills', '')
            ])
        
        # Return downloadable CSV
        filename = f'cluster_export_{timestamp}.csv'
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

@cluster_bp.route('/bulk_move_candidates', methods=['POST'])
@cluster_required
def bulk_move_candidates():
    from bson import ObjectId
    from models_mongo import candidates_collection, ActivityLog
    
    try:
        data = request.get_json()
        candidate_ids = data.get('candidate_ids', [])
        new_status = data.get('new_status', '')
        
        if not candidate_ids:
            return jsonify({'success': False, 'message': 'No candidates selected'})
        
        if not new_status:
            return jsonify({'success': False, 'message': 'No status specified'})
        
        # Convert string IDs to ObjectId
        object_ids = [ObjectId(cid) for cid in candidate_ids]
        
        # Get candidate details for logging
        candidates = list(candidates_collection.find({'_id': {'$in': object_ids}}))
        
        # Update candidates status
        result = candidates_collection.update_many(
            {'_id': {'$in': object_ids}},
            {'$set': {'status': new_status}}
        )
        
        if result.modified_count > 0:
            # Log the bulk move activity
            activity = ActivityLog(
                user_email=current_user.email,
                action=f'Bulk moved candidates to {new_status}',
                target_email='',
                details=f'Cluster moved {result.modified_count} candidates to {new_status}: {", ".join([c.get("first_name", "") + " " + c.get("last_name", "") for c in candidates])}'
            )
            activity.save()
            
            return jsonify({'success': True, 'message': f'Successfully moved {result.modified_count} candidates to {new_status}'})
        else:
            return jsonify({'success': False, 'message': 'No candidates found to update'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error moving candidates: {str(e)}'}) 

@cluster_bp.route('/logout')
@cluster_required
def logout():
    """Cluster logout route - redirects to home page"""
    from flask_login import logout_user
    logout_user()
    return redirect(url_for('index'))

@cluster_bp.route('/get-clusters')
@cluster_required
def get_clusters():
    """Get all available clusters for filtering"""
    from models_mongo import users_collection
    
    # Get unique clusters from users
    cluster_users = list(users_collection.find({'cluster': {'$exists': True, '$ne': None}}))
    cluster_names = []
    for user in cluster_users:
        if user.get('cluster') and user['cluster'] not in cluster_names:
            cluster_names.append(user['cluster'])
    
    # If no clusters found, provide default options
    if not cluster_names:
        cluster_names = ['Tech Cluster', 'Sales Cluster', 'Marketing Cluster', 'Operations Cluster']
    
    return jsonify({
        'clusters': cluster_names,
        'total': len(cluster_names)
    })

@cluster_bp.route('/get_candidates')
@cluster_required
def get_candidates():
    """Get candidates data for the cluster dashboard"""
    from models_mongo import candidates_collection, users_collection
    
    try:
        # Get filter parameters
        status_filter = request.args.get('status')
        hr_filter = request.args.get('hr')
        manager_filter = request.args.get('manager')
        cluster_filter = request.args.get('cluster')
        
        # Build query based on filters
        query = {}
        if status_filter and status_filter != 'All Statuses':
            query['status'] = status_filter
        if hr_filter and hr_filter != 'All HRs':
            hr_user = users_collection.find_one({'role': 'hr_role', 'name': hr_filter, 'is_active': True})
            if hr_user:
                query['assigned_by'] = hr_user['email']
        if manager_filter and manager_filter != 'All Managers':
            manager_user = users_collection.find_one({'role': 'manager', 'name': manager_filter, 'is_active': True})
            if manager_user:
                query['manager_email'] = manager_user['email']
        if cluster_filter and cluster_filter != 'All Clusters':
            cluster_users = list(users_collection.find({'cluster': cluster_filter, 'is_active': True}))
            if cluster_users:
                cluster_emails = [user['email'] for user in cluster_users]
                query['$or'] = [
                    {'assigned_by': {'$in': cluster_emails}},
                    {'manager_email': {'$in': cluster_emails}}
                ]
        
        # Get candidates
        all_candidates = list(candidates_collection.find(query))
        
        # Calculate statistics
        total_candidates = len(all_candidates)
        selected_candidates = len([c for c in all_candidates if c.get('status') == 'Selected'])
        not_selected_candidates = len([c for c in all_candidates if c.get('status') == 'Not Selected'])
        pending_candidates = len([c for c in all_candidates if c.get('status') in ['Assigned', 'Pending']])
        
        # Calculate onboarding statistics
        onboarded_candidates = len([c for c in all_candidates if c.get('onboarding_status') == 'Onboarded'])
        not_onboarded_candidates = len([c for c in all_candidates if c.get('status') == 'Selected' and c.get('onboarding_status') != 'Onboarded'])
        
        # Convert ObjectId to string and fix data for frontend
        for candidate in all_candidates:
            if '_id' in candidate:
                candidate['_id'] = str(candidate['_id'])
            if 'created_at' in candidate and candidate['created_at']:
                # Handle both datetime objects and strings
                if hasattr(candidate['created_at'], 'isoformat'):
                    candidate['created_at'] = candidate['created_at'].isoformat()
                # If it's already a string, keep it as is
            
            # Fix name display
            first_name = candidate.get('first_name', '')
            last_name = candidate.get('last_name', '')
            candidate['name'] = f"{first_name} {last_name}".strip() or 'Unknown'
            
            # Map date of birth field for frontend compatibility
            if candidate.get('dob'):
                candidate['date_of_birth'] = candidate['dob']
            else:
                candidate['date_of_birth'] = None
            
            # Get Recruiter name from email (since candidates are uploaded by recruiters)
            recruiter_email = candidate.get('assigned_by')
            if recruiter_email:
                recruiter_user = users_collection.find_one({'email': recruiter_email, 'role': 'recruiter'})
                if recruiter_user:
                    candidate['recruiter_name'] = recruiter_user.get('name', 'Unknown Recruiter')
                else:
                    # Fallback: check if it's an HR user
                    hr_user = users_collection.find_one({'email': recruiter_email, 'role': 'hr_role'})
                    if hr_user:
                        candidate['recruiter_name'] = hr_user.get('name', 'Unknown HR User')
                    else:
                        candidate['recruiter_name'] = 'Unknown User'
            else:
                candidate['recruiter_name'] = None  # Don't show recruiter field if no one assigned
            
            # Get Manager name from email
            manager_email = candidate.get('manager_email')
            if manager_email:
                manager_user = users_collection.find_one({'email': manager_email, 'role': 'manager'})
                candidate['manager_name'] = manager_user.get('name', 'Unknown Manager') if manager_user else 'Unknown Manager'
            else:
                candidate['manager_name'] = None  # Don't show Manager field if no manager assigned
            
            # Get cluster info
            cluster_name = None
            if recruiter_email:
                recruiter_user = users_collection.find_one({'email': recruiter_email})
                if recruiter_user and recruiter_user.get('cluster'):
                    cluster_name = recruiter_user['cluster']
            elif manager_email:
                manager_user = users_collection.find_one({'email': manager_email})
                if manager_user and manager_user.get('cluster'):
                    cluster_name = manager_user['cluster']
            candidate['cluster_name'] = cluster_name if cluster_name else None
            
            # Add onboarding status information
            candidate['onboarding_status'] = candidate.get('onboarding_status', 'Not Started')
            candidate['onboarding_date'] = candidate.get('onboarding_date', None)
            candidate['onboarding_notes'] = candidate.get('onboarding_notes', '')
        
        return jsonify({
            'success': True,
            'total_candidates': total_candidates,
            'selected_candidates': selected_candidates,
            'not_selected_candidates': not_selected_candidates,
            'pending_candidates': pending_candidates,
            'onboarded_candidates': onboarded_candidates,
            'not_onboarded_candidates': not_onboarded_candidates,
            'candidates': all_candidates[:50],  # Limit to first 50 for performance
            'total_returned': min(50, total_candidates)
        })
        
    except Exception as e:
        print(f"Error getting candidates: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'total_candidates': 0,
            'selected_candidates': 0,
            'not_selected_candidates': 0,
            'pending_candidates': 0,
            'onboarded_candidates': 0,
            'not_onboarded_candidates': 0,
            'candidates': []
        }), 500

@cluster_bp.route('/get_filter_data')
@cluster_required  
def get_filter_data_underscore():
    """Get all filter data (HR, Managers, Clusters) for dynamic updates - underscore version"""
    return get_filter_data()

@cluster_bp.route('/get-filter-data')
@cluster_required
def get_filter_data():
    """Get all filter data (HR, Managers, Clusters) for dynamic updates"""
    from models_mongo import users_collection
    
    try:
        # Get HR names for filter (sorted alphabetically)
        hr_users = list(users_collection.find({'role': 'hr_role', 'is_active': True}).sort('name', 1))
        hr_names = []
        for hr in hr_users:
            if hr.get('name') and hr['name'].strip():
                hr_names.append(hr['name'].strip())
        
        # Get Manager names for filter (sorted alphabetically)
        manager_users = list(users_collection.find({'role': 'manager', 'is_active': True}).sort('name', 1))
        manager_names = []
        for manager in manager_users:
            if manager.get('name') and manager['name'].strip():
                manager_names.append(manager['name'].strip())
        
        # Get unique clusters from users (sorted alphabetically)
        cluster_users = list(users_collection.find({'cluster': {'$exists': True, '$ne': None}}).sort('cluster', 1))
        cluster_names = []
        for user in cluster_users:
            if user.get('cluster') and user['cluster'].strip() and user['cluster'] not in cluster_names:
                cluster_names.append(user['cluster'].strip())
        
        # If no clusters found, provide default options
        if not cluster_names:
            cluster_names = ['Tech Cluster', 'Sales Cluster', 'Marketing Cluster', 'Operations Cluster']
        
        # Add timestamp for debugging
        from datetime import datetime
        timestamp = datetime.now().isoformat()
        
        return jsonify({
            'hr_names': hr_names,
            'manager_names': manager_names,
            'cluster_names': cluster_names,
            'total_hr': len(hr_names),
            'total_managers': len(manager_names),
            'total_clusters': len(cluster_names),
            'last_updated': timestamp,
            'success': True
        })
        
    except Exception as e:
        print(f"Error getting filter data: {e}")
        return jsonify({
            'error': str(e),
            'success': False,
            'hr_names': [],
            'manager_names': [],
            'cluster_names': []
        }), 500

@cluster_bp.route('/get-chart-data')
@cluster_required
def get_chart_data():
    """Get live chart data for 3D visualization"""
    from models_mongo import candidates_collection, users_collection
    try:
        # Get filter parameters
        manager_filter = request.args.get('manager', '')
        hr_filter = request.args.get('hr', '')
        status_filter = request.args.get('status', '')
        
        # Build query based on filters
        query = {}
        if manager_filter:
            query['manager_email'] = manager_filter
        if hr_filter:
            query['hr_email'] = hr_filter
        if status_filter:
            query['status'] = status_filter
        
        # Get all candidates (or filtered)
        all_candidates = list(candidates_collection.find(query))
        
        # Debug logging
        print(f"Total candidates found: {len(all_candidates)}")
        if all_candidates:
            print(f"Sample candidate fields: {list(all_candidates[0].keys())}")
        
        # Manager Performance Data
        manager_data = []
        if not manager_filter:
            # Get all managers
            managers = list(users_collection.find({'role': 'manager', 'is_active': True}).sort('name', 1))
            print(f"Found {len(managers)} managers")
            for manager in managers:
                manager_email = manager.get('email', '')
                # Try different possible field names for manager assignment
                manager_candidates = [c for c in all_candidates if 
                                   c.get('manager_email') == manager_email or 
                                   c.get('assigned_to') == manager_email or
                                   c.get('manager') == manager_email]
                
                selected_count = len([c for c in manager_candidates if c.get('status') == 'Selected'])
                not_selected_count = len([c for c in manager_candidates if c.get('status') == 'Not Selected'])
                pending_count = len([c for c in manager_candidates if c.get('status') == 'Pending'])
                
                manager_data.append({
                    'manager_name': manager.get('name', 'Unknown'),
                    'selected_count': selected_count,
                    'not_selected_count': not_selected_count,
                    'pending_count': pending_count
                })
        else:
            # Single manager data
            manager_candidates = [c for c in all_candidates if 
                               c.get('manager_email') == manager_filter or 
                               c.get('assigned_to') == manager_filter or
                               c.get('manager') == manager_filter]
            selected_count = len([c for c in manager_candidates if c.get('status') == 'Selected'])
            not_selected_count = len([c for c in manager_candidates if c.get('status') == 'Not Selected'])
            pending_count = len([c for c in manager_candidates if c.get('status') == 'Pending'])
            
            manager_data.append({
                'manager_name': manager_filter,
                'selected_count': selected_count,
                'not_selected_count': not_selected_count,
                'pending_count': pending_count
            })
        
        # HR Performance Data
        hr_data = []
        if not hr_filter:
            # Get all HR users
            hr_users = list(users_collection.find({'role': 'hr_role', 'is_active': True}).sort('name', 1))
            print(f"Found {len(hr_users)} HR users")
            for hr_user in hr_users:
                hr_email = hr_user.get('email', '')
                # Try different possible field names for HR assignment
                hr_candidates = [c for c in all_candidates if 
                               c.get('hr_email') == hr_email or 
                               c.get('assigned_by') == hr_email or
                               c.get('hr') == hr_email]
                
                hr_data.append({
                    'hr_name': hr_user.get('name', 'Unknown'),
                    'candidate_count': len(hr_candidates)
                })
        else:
            # Single HR data
            hr_candidates = [c for c in all_candidates if 
                           c.get('hr_email') == hr_filter or 
                           c.get('assigned_by') == hr_filter or
                           c.get('hr') == hr_filter]
            
            hr_data.append({
                'hr_name': hr_filter,
                'candidate_count': len(hr_candidates)
            })
        
        # Status Distribution Data
        status_data = {}
        if not status_filter:
            statuses = ['Selected', 'Not Selected', 'Pending', 'Assigned', 'Reassigned']
            for status in statuses:
                count = len([c for c in all_candidates if c.get('status') == status])
                if count > 0:  # Only include statuses with candidates
                    status_data[status] = count
        else:
            # Single status data
            status_data[status_filter] = len(all_candidates)
        
        from datetime import datetime
        
        # If no real data found, provide sample data for demonstration
        if not manager_data and not hr_data and not status_data:
            print("No real data found, providing sample data")
            manager_data = [
                {'manager_name': 'Manager A', 'selected_count': 12, 'not_selected_count': 3, 'pending_count': 2},
                {'manager_name': 'Manager B', 'selected_count': 19, 'not_selected_count': 5, 'pending_count': 3},
                {'manager_name': 'Manager C', 'selected_count': 8, 'not_selected_count': 2, 'pending_count': 1},
                {'manager_name': 'Manager D', 'selected_count': 15, 'not_selected_count': 4, 'pending_count': 2}
            ]
            hr_data = [
                {'hr_name': 'HR A', 'candidate_count': 30},
                {'hr_name': 'HR B', 'candidate_count': 25},
                {'hr_name': 'HR C', 'candidate_count': 20},
                {'hr_name': 'HR D', 'candidate_count': 25}
            ]
            status_data = {
                'Selected': 45,
                'Not Selected': 25,
                'Pending': 15,
                'Assigned': 10,
                'Reassigned': 5
            }
        
        print(f"Returning data: {len(manager_data)} managers, {len(hr_data)} HRs, {len(status_data)} statuses")
        
        return jsonify({
            'success': True,
            'managerData': manager_data,
            'hrData': hr_data,
            'statusData': status_data,
            'totalCandidates': len(all_candidates),
            'lastUpdated': datetime.now().isoformat()
        })
        
    except Exception as e:
        print(f"Error getting chart data: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'managerData': [],
            'hrData': [],
            'statusData': {},
            'totalCandidates': 0
        }), 500

@cluster_bp.route('/analytics/<tab_name>')
@cluster_required
def get_analytics_data(tab_name):
    """Get analytics data for Cluster dashboard"""
    try:
        from models_mongo import candidates_collection, users_collection
        from collections import defaultdict
        
        # Get all candidates across all clusters
        candidates = list(candidates_collection.find({}))
        
        # Process data based on tab requested
        if tab_name == 'overview':
            # Status distribution across all clusters
            status_counts = defaultdict(int)
            for candidate in candidates:
                status = candidate.get('status', 'Pending')
                status_counts[status] += 1
            
            # Cluster distribution
            cluster_data = defaultdict(int)
            hr_users = list(users_collection.find({'role': 'hr'}))
            manager_users = list(users_collection.find({'role': 'manager'}))
            
            for candidate in candidates:
                hr_email = candidate.get('assigned_by')
                manager_email = candidate.get('manager_email')
                
                # Try to get cluster from HR
                if hr_email:
                    hr_user = next((u for u in hr_users if u.get('email') == hr_email), None)
                    if hr_user and hr_user.get('cluster'):
                        cluster_data[hr_user['cluster']] += 1
                        continue
                
                # Try to get cluster from Manager
                if manager_email:
                    manager_user = next((u for u in manager_users if u.get('email') == manager_email), None)
                    if manager_user and manager_user.get('cluster'):
                        cluster_data[manager_user['cluster']] += 1
                        continue
                
                # Default cluster
                cluster_data['General'] += 1
            
            return jsonify({
                'success': True,
                'status_counts': dict(status_counts),
                'cluster_data': dict(cluster_data),
                'total_candidates': len(candidates),
                'total_hrs': len(hr_users),
                'total_managers': len(manager_users),
                'total_clusters': len(set(cluster_data.keys())),
                'last_updated': datetime.now().isoformat()
            })
            
        elif tab_name == 'trends':
            # Trend data over time across all clusters
            trend_data = []
            now = datetime.now()
            for i in range(30, 0, -5):  # Last 30 days in 5-day intervals
                date = now - timedelta(days=i)
                date_str = date.strftime('%m/%d')
                
                # Count applications and selections for this period
                applications = 0
                selections = 0
                for candidate in candidates:
                    created_date = candidate.get('created_at', now)
                    if isinstance(created_date, str):
                        try:
                            created_date = datetime.fromisoformat(created_date.replace('Z', '+00:00'))
                        except:
                            created_date = now
                    
                    if abs((created_date - date).days) <= 2:  # Within 2 days
                        applications += 1
                        if candidate.get('status') == 'Selected':
                            selections += 1
                
                success_rate = (selections / applications * 100) if applications > 0 else 0
                
                trend_data.append({
                    'date': date_str,
                    'applications': applications,
                    'selections': selections,
                    'success_rate': round(success_rate, 1)
                })
            
            return jsonify({
                'success': True,
                'trend_data': trend_data,
                'last_updated': datetime.now().isoformat()
            })
            
        elif tab_name == 'performance':
            # Multi-cluster performance comparison
            cluster_performance = defaultdict(lambda: {'candidates': 0, 'selected': 0, 'pending': 0})
            hr_users = list(users_collection.find({'role': 'hr'}))
            manager_users = list(users_collection.find({'role': 'manager'}))
            
            for candidate in candidates:
                cluster_name = 'General'
                
                # Get cluster from HR or Manager
                hr_email = candidate.get('assigned_by')
                manager_email = candidate.get('manager_email')
                
                if hr_email:
                    hr_user = next((u for u in hr_users if u.get('email') == hr_email), None)
                    if hr_user and hr_user.get('cluster'):
                        cluster_name = hr_user['cluster']
                elif manager_email:
                    manager_user = next((u for u in manager_users if u.get('email') == manager_email), None)
                    if manager_user and manager_user.get('cluster'):
                        cluster_name = manager_user['cluster']
                
                cluster_performance[cluster_name]['candidates'] += 1
                if candidate.get('status') == 'Selected':
                    cluster_performance[cluster_name]['selected'] += 1
                elif candidate.get('status') == 'Pending':
                    cluster_performance[cluster_name]['pending'] += 1
            
            return jsonify({
                'success': True,
                'cluster_performance': dict(cluster_performance),
                'last_updated': datetime.now().isoformat()
            })
            
        elif tab_name == 'reports':
            # Comprehensive reporting data
            # HR and Manager distribution by cluster
            hr_users = list(users_collection.find({'role': 'hr'}))
            manager_users = list(users_collection.find({'role': 'manager'}))
            
            hr_by_cluster = defaultdict(int)
            manager_by_cluster = defaultdict(int)
            
            for hr in hr_users:
                cluster = hr.get('cluster', 'General')
                hr_by_cluster[cluster] += 1
            
            for manager in manager_users:
                cluster = manager.get('cluster', 'General')
                manager_by_cluster[cluster] += 1
            
            # Response time analysis (mock data)
            response_times = {
                'HR Response': 2.3,
                'Manager Review': 1.8,
                'Final Decision': 3.2,
                'Overall Average': 2.4
            }
            
            return jsonify({
                'success': True,
                'hr_by_cluster': dict(hr_by_cluster),
                'manager_by_cluster': dict(manager_by_cluster),
                'response_times': response_times,
                'last_updated': datetime.now().isoformat()
            })
        
        else:
            return jsonify({
                'success': False,
                'message': f'Unknown tab: {tab_name}'
            }), 400
            
    except Exception as e:
        print(f"Error in Cluster analytics {tab_name}: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to load analytics data'
        }), 500

@cluster_bp.route('/hr_performance')
@cluster_required
def get_hr_performance():
    """Get HR personnel performance data"""
    try:
        from models_mongo import candidates_collection, users_collection
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # Get all HR users
        hr_users = list(users_collection.find({'role': 'hr_role', 'is_active': True}))
        hr_performance = []
        
        for hr in hr_users:
            hr_email = hr.get('email')
            
            # Get candidates assigned by this HR
            assigned_candidates = list(candidates_collection.find({'assigned_by': hr_email}))
            
            # Calculate metrics
            total_assigned = len(assigned_candidates)
            pending = len([c for c in assigned_candidates if c.get('status') == 'Pending'])
            selected = len([c for c in assigned_candidates if c.get('status') == 'Selected'])
            not_selected = len([c for c in assigned_candidates if c.get('status') == 'Not Selected'])
            
            # Calculate last activity (most recent candidate assignment)
            last_activity = 'N/A'
            if assigned_candidates:
                latest = max(assigned_candidates, key=lambda x: x.get('created_at', datetime.min))
                if latest.get('created_at'):
                    try:
                        if isinstance(latest['created_at'], str):
                            last_date = datetime.fromisoformat(latest['created_at'].replace('Z', '+00:00'))
                        else:
                            last_date = latest['created_at']
                        last_activity = last_date.strftime('%m/%d/%Y')
                    except:
                        last_activity = 'Recent'
            
            hr_performance.append({
                'name': hr.get('first_name', '') + ' ' + hr.get('last_name', ''),
                'email': hr_email,
                'cluster': hr.get('cluster', 'General'),
                'total_assigned': total_assigned,
                'pending': pending,
                'selected': selected,
                'not_selected': not_selected,
                'last_activity': last_activity,
                'avg_response_time': '2.5 hours'  # Mock data for now
            })
        
        return jsonify({
            'success': True,
            'hr_personnel': hr_performance
        })
        
    except Exception as e:
        print(f"Error getting HR performance: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to load HR performance data'
        }), 500

@cluster_bp.route('/recruiter_performance')
@cluster_required
def get_recruiter_performance():
    """Get Recruiter personnel performance data"""
    try:
        from models_mongo import candidates_collection, users_collection
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # Get all Recruiter users
        recruiter_users = list(users_collection.find({'role': 'recruiter'}))
        recruiter_performance = []
        
        for recruiter in recruiter_users:
            recruiter_email = recruiter.get('email')
            
            # Get candidates created/uploaded by this recruiter
            uploaded_candidates = list(candidates_collection.find({'assigned_by': recruiter_email}))
            
            # Calculate metrics
            total_uploaded = len(uploaded_candidates)
            pending = len([c for c in uploaded_candidates if c.get('status') == 'Pending'])
            assigned = len([c for c in uploaded_candidates if c.get('status') == 'Assigned'])
            selected = len([c for c in uploaded_candidates if c.get('status') == 'Selected'])
            
            # More robust detection of "not selected" candidates
            not_selected = len([c for c in uploaded_candidates if 
                c.get('status') in ['Not Selected', 'Rejected', 'Declined', 'Failed', 'Not Approved']
            ])
            
            # Calculate last activity (most recent candidate upload)
            last_activity = 'N/A'
            if uploaded_candidates:
                latest = max(uploaded_candidates, key=lambda x: x.get('created_at', datetime.min))
                if latest.get('created_at'):
                    try:
                        if isinstance(latest['created_at'], str):
                            last_date = datetime.fromisoformat(latest['created_at'].replace('Z', '+00:00'))
                        else:
                            last_date = latest['created_at']
                        last_activity = last_date.strftime('%m/%d/%Y')
                    except:
                        last_activity = 'Recent'
            
            recruiter_performance.append({
                'name': recruiter.get('name', 'Unknown'),
                'email': recruiter_email,
                'cluster': recruiter.get('cluster', 'General'),
                'total_uploaded': total_uploaded,
                'pending': pending,
                'assigned': assigned,
                'selected': selected,
                'not_selected': not_selected,
                'last_activity': last_activity,
                'avg_response_time': '1.2 hours'  # Mock data for now
            })
        
        return jsonify({
            'success': True,
            'recruiter_personnel': recruiter_performance
        })
        
    except Exception as e:
        print(f"Error getting Recruiter performance: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to load Recruiter performance data'
        }), 500

@cluster_bp.route('/manager_performance')
@cluster_required
def get_manager_performance():
    """Get Manager personnel performance data"""
    try:
        from models_mongo import candidates_collection, users_collection
        from collections import defaultdict
        from datetime import datetime, timedelta
        
        # Get all Manager users
        manager_users = list(users_collection.find({'role': 'manager'}))
        manager_performance = []
        
        for manager in manager_users:
            manager_email = manager.get('email')
            
            # Get candidates assigned to this manager
            assigned_candidates = list(candidates_collection.find({'manager_email': manager_email}))
            
            # Calculate metrics
            total_reviewed = len([c for c in assigned_candidates if c.get('status') in ['Selected', 'Not Selected', 'Rejected', 'Declined', 'Failed', 'Not Approved']])
            # Count reassigned candidates (those with reassigned_by_manager field)
            pending_review = len([c for c in assigned_candidates if c.get('reassigned_by_manager') is not None])
            selected = len([c for c in assigned_candidates if c.get('status') == 'Selected'])
            
            # More robust detection of "not selected" candidates
            not_selected = len([c for c in assigned_candidates if 
                c.get('status') in ['Not Selected', 'Rejected', 'Declined', 'Failed', 'Not Approved']
            ])
            
            # Calculate last review activity
            last_review = 'N/A'
            reviewed_candidates = [c for c in assigned_candidates if c.get('status') in ['Selected', 'Not Selected', 'Rejected', 'Declined', 'Failed', 'Not Approved']]
            if reviewed_candidates:
                # For now, use created_at as proxy for review date
                latest = max(reviewed_candidates, key=lambda x: x.get('created_at', datetime.min))
                if latest.get('created_at'):
                    try:
                        if isinstance(latest['created_at'], str):
                            last_date = datetime.fromisoformat(latest['created_at'].replace('Z', '+00:00'))
                        else:
                            last_date = latest['created_at']
                        last_review = last_date.strftime('%m/%d/%Y')
                    except:
                        last_review = 'Recent'
            
            manager_performance.append({
                'name': manager.get('first_name', '') + ' ' + manager.get('last_name', ''),
                'email': manager_email,
                'cluster': manager.get('cluster', 'General'),
                'total_reviewed': total_reviewed,
                'pending_review': pending_review,
                'selected': selected,
                'not_selected': not_selected,
                'last_review': last_review,
                'avg_review_time': '1.8 hours'  # Mock data for now
            })
        
        return jsonify({
            'success': True,
            'manager_personnel': manager_performance
        })
        
    except Exception as e:
        print(f"Error getting Manager performance: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to load Manager performance data'
        }), 500

@cluster_bp.route('/weekly_trends')
@cluster_required
def get_weekly_trends():
    """Get weekly performance trends"""
    try:
        from models_mongo import candidates_collection
        from datetime import datetime, timedelta
        
        # Get current week data
        now = datetime.now()
        week_start = now - timedelta(days=7)
        
        # Get candidates from this week
        week_candidates = list(candidates_collection.find({
            'created_at': {'$gte': week_start.isoformat()}
        }))
        
        # Calculate metrics
        assigned = len(week_candidates)
        pending = len([c for c in week_candidates if c.get('status') == 'Pending'])
        selected = len([c for c in week_candidates if c.get('status') == 'Selected'])
        
        # More robust detection of rejected candidates
        rejected = len([c for c in week_candidates if 
            c.get('status') in ['Not Selected', 'Rejected', 'Declined', 'Failed', 'Not Approved']
        ])
        
        # Mock trend data (would normally compare with previous week)
        trends = {
            'assigned': assigned,
            'pending': pending,
            'selected': selected,
            'rejected': rejected,
            'assigned_trend': '↗ +12%',
            'pending_trend': '↘ -8%',
            'selected_trend': '↗ +25%',
            'rejected_trend': '↘ -15%'
        }
        
        return jsonify({
            'success': True,
            **trends
        })
        
    except Exception as e:
        print(f"Error getting weekly trends: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to load weekly trends'
        }), 500 